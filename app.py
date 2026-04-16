import os
import sys
import base64
import webview               # <-- ADICIONADO: Webview
from threading import Thread # <-- ADICIONADO: Thread
from io import BytesIO
from html import escape
from datetime import datetime, date
from functools import wraps

from dotenv import load_dotenv

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    send_from_directory,
)
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, case
from werkzeug.security import generate_password_hash, check_password_hash

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    Image,
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    KeepTogether,
)

from PIL import Image as PILImage # <-- ADICIONADO: Biblioteca para comprimir imagens

# =========================
# FUNÇÃO PARA O PYINSTALLER
# =========================
def resource_path(relative_path):
    """ Retorna o caminho absoluto para o recurso, compatível com o PyInstaller """
    try:
        # O PyInstaller extrai os ficheiros para esta pasta temporária
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


app = Flask(__name__, 
            template_folder=resource_path('templates'),
            static_folder=resource_path('static'))
app.config["SECRET_KEY"] = "dev-secret-change-me"
app.config["TEMPLATES_AUTO_RELOAD"] = True

# =========================
# CONFIGURAÇÃO ORACLE DB SEGURA (.ENV EMBUTIDO)
# =========================
dotenv_path = resource_path('.env')
load_dotenv(dotenv_path)

DB_USER = os.getenv("DB_USER")
DB_PASS = os.getenv("DB_PASS")
DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT", "1521")
DB_SERVICE = os.getenv("DB_SERVICE")

if not all([DB_USER, DB_PASS, DB_HOST, DB_SERVICE]):
    raise ValueError("Variáveis de ambiente do banco de dados não encontradas. Verifique o arquivo .env!")

app.config["SQLALCHEMY_DATABASE_URI"] = f"oracle+oracledb://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/?service_name={DB_SERVICE}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

# A pasta de uploads DEVE ficar sempre ao lado do .exe, nunca na pasta temporária
if getattr(sys, 'frozen', False):
    exe_dir = os.path.dirname(sys.executable)
else:
    exe_dir = os.path.dirname(os.path.abspath(__file__))

# <-- ALTERAÇÃO: Removido a criação física da pasta uploads já que usamos Base64 no banco -->
# UPLOAD_FOLDER = os.path.join(exe_dir, "uploads")
# os.makedirs(UPLOAD_FOLDER, exist_ok=True)
# app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

ALLOWED_IMAGE_EXTENSIONS = {"png", "jpg", "jpeg", "webp"}

db = SQLAlchemy(app)


# =========================
# CONSTANTES / CONTROLE DE VERSÃO
# =========================
APP_NAME = "LIVRO_OCORRENCIAS"
APP_VERSION = "1.0.2"  # <- Altere este número sempre que compilar um novo executável

TURNOS_VALIDOS = {"TURNO A", "TURNO B", "TURNO C", "ADM"}
TIPOS_VALIDOS = {
    "ROTINA",
    "DESVIO DE PROCESSO",
    "EXTRAVIO",
    "INCIDENTE",
    "MANUTENCAO",
    "MANUTENÇÃO",
    "PENDENCIA",
    "PENDÊNCIA",
    "INFORMATIVO",
}
PRIORIDADES_VALIDAS = {"BAIXA", "MEDIA", "MÉDIA", "ALTA", "CRITICA", "CRÍTICA"}
STATUS_VALIDOS = {"EM ABERTO", "EM ACOMPANHAMENTO", "FINALIZADO"}


# =========================
# MODELOS
# =========================
class SistemaConfig(db.Model):
    __tablename__ = "SISTEMA_CONFIG"
    
    id = db.Column("ID", db.Integer, primary_key=True)
    versao_exigida = db.Column("VERSAO_EXIGIDA", db.String(20), nullable=True) 
    versao_livro = db.Column("VERSAO_LIVRO", db.String(20), nullable=True)     


class Site(db.Model):
    __tablename__ = "SITES"
    
    id_site = db.Column("ID_SITE", db.Integer, primary_key=True)
    nome_site = db.Column("NOME_SITE", db.String(100), nullable=False)


class User(db.Model):
    __tablename__ = "users_livro"

    id = db.Column(db.Integer, db.Identity(start=1), primary_key=True)
    nome = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="USER")
    site = db.Column(db.String(80), nullable=True)
    is_active = db.Column(db.Boolean, nullable=False, default=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.now)

    def set_password(self, senha: str):
        self.password_hash = generate_password_hash(senha)

    def check_password(self, senha: str) -> bool:
        return check_password_hash(self.password_hash, senha)


class OcorrenciaTurno(db.Model):
    __tablename__ = "ocorrencias_turno"

    id = db.Column(db.Integer, db.Identity(start=1), primary_key=True)
    data_ocorrencia = db.Column(db.Date, nullable=False, default=date.today)
    data_hora_registro = db.Column(db.DateTime, nullable=False, default=datetime.now)

    site = db.Column(db.String(80), nullable=False)
    turno = db.Column(db.String(30), nullable=False)
    setor = db.Column(db.String(100), nullable=False)
    tipo_ocorrencia = db.Column(db.String(100), nullable=False)
    prioridade = db.Column(db.String(20), nullable=False)

    responsavel_saida = db.Column(db.String(120), nullable=False)
    responsavel_entrada = db.Column(db.String(120), nullable=False)

    descricao = db.Column(db.Text, nullable=False)
    efetivo = db.Column(db.Text, nullable=False)

    assinatura_saida = db.Column(db.Text, nullable=True)
    assinatura_entrada = db.Column(db.Text, nullable=True)

    # <-- ALTERADO: String(255) alterado para Text (CLOB no Oracle) para suportar as imagens base64 -->
    imagem_1 = db.Column(db.Text, nullable=True)
    imagem_2 = db.Column(db.Text, nullable=True)
    imagem_3 = db.Column(db.Text, nullable=True)
    imagem_4 = db.Column(db.Text, nullable=True)

    acoes_tomadas = db.Column(db.Text, nullable=True)
    pendencias = db.Column(db.Text, nullable=True)

    status = db.Column(db.String(40), nullable=False)

    criado_por = db.Column(db.String(120), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.now)
    updated_at = db.Column(db.DateTime, nullable=True)

    def to_dict(self):
        return {
            "id": self.id,
            "data_ocorrencia": self.data_ocorrencia.strftime("%d/%m/%Y") if self.data_ocorrencia else "",
            "data_hora_registro": self.data_hora_registro.strftime("%d/%m/%Y %H:%M")
            if self.data_hora_registro
            else "",
            "site": self.site,
            "turno": self.turno,
            "setor": self.setor,
            "tipo_ocorrencia": self.tipo_ocorrencia,
            "prioridade": self.prioridade,
            "responsavel_saida": self.responsavel_saida,
            "responsavel_entrada": self.responsavel_entrada,
            "descricao": self.descricao,
            "efetivo": self.efetivo or "",
            "assinatura_saida": self.assinatura_saida or "",
            "assinatura_entrada": self.assinatura_entrada or "",
            "imagem_1": self.imagem_1 or "",
            "imagem_2": self.imagem_2 or "",
            "imagem_3": self.imagem_3 or "",
            "imagem_4": self.imagem_4 or "",
            "acoes_tomadas": self.acoes_tomadas or "",
            "pendencias": self.pendencias or "",
            "status": self.status,
            "criado_por": self.criado_por or "",
            "criado_em": self.created_at.strftime("%d/%m/%Y %H:%M") if self.created_at else "",
            "atualizado_em": self.updated_at.strftime("%d/%m/%Y %H:%M") if self.updated_at else "",
        }


def garantir_colunas_ocorrencias():
    try:
        with db.engine.connect() as conn:
            result = conn.execute(db.text("SELECT lower(column_name) FROM user_tab_columns WHERE table_name = 'OCORRENCIAS_TURNO'"))
            colunas = {row[0] for row in result.fetchall()}

            if not colunas:
                return

            comandos = []

            if "efetivo" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD efetivo CLOB")

            if "assinatura_saida" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD assinatura_saida CLOB")

            if "assinatura_entrada" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD assinatura_entrada CLOB")

            # <-- ALTERADO: Colunas de imagem geradas como CLOB direto caso a tabela seja nova -->
            if "imagem_1" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD imagem_1 CLOB")

            if "imagem_2" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD imagem_2 CLOB")

            if "imagem_3" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD imagem_3 CLOB")

            if "imagem_4" not in colunas:
                comandos.append("ALTER TABLE ocorrencias_turno ADD imagem_4 CLOB")

            for sql in comandos:
                conn.execute(db.text(sql))

            conn.commit()
    except Exception as e:
        print(f"Aviso ao verificar colunas no Oracle (garantir_colunas): {e}")


def garantir_colunas_sistema_config():
    try:
        with db.engine.connect() as conn:
            result = conn.execute(db.text("SELECT lower(column_name) FROM user_tab_columns WHERE table_name = 'SISTEMA_CONFIG'"))
            colunas = {row[0] for row in result.fetchall()}
            
            if colunas and "versao_livro" not in colunas:
                conn.execute(db.text("ALTER TABLE SISTEMA_CONFIG ADD VERSAO_LIVRO VARCHAR2(20)"))
                conn.execute(db.text(f"UPDATE SISTEMA_CONFIG SET VERSAO_LIVRO = '{APP_VERSION}'"))
                conn.commit()
                print("Coluna VERSAO_LIVRO adicionada à SISTEMA_CONFIG com sucesso.")
    except Exception as e:
        print(f"Aviso ao verificar SISTEMA_CONFIG: {e}")


# =========================
# MIDDLEWARE E SEGURANÇA
# =========================
@app.before_request
def check_app_version():
    rotas_livres = ('static', 'serve_upload', 'versao_invalida', 'criar_banco')
    if request.endpoint in rotas_livres:
        return

    try:
        config_db = SistemaConfig.query.first()
        
        if config_db and config_db.versao_livro and config_db.versao_livro != APP_VERSION:
            return redirect(url_for('versao_invalida'))
    except Exception:
        pass


# <-- ADICIONADO: Marca d'água persistente via HTML -->
@app.after_request
def add_watermark(response):
    if response.content_type and response.content_type.startswith('text/html'):
        watermark = b'<div style="position: fixed; bottom: 15px; right: 15px; opacity: 0.6; font-family: Arial, sans-serif; font-size: 13px; font-weight: bold; color: #888; z-index: 9999; pointer-events: none;">Powered by Security</div></body>'
        response.data = response.data.replace(b'</body>', watermark)
    return response


def login_required(funcao):
    @wraps(funcao)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar o sistema.", "warning")
            return redirect(url_for("login"))
        return funcao(*args, **kwargs)
    return wrapper


def admin_required(funcao):
    @wraps(funcao)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            flash("Faça login para acessar o sistema.", "warning")
            return redirect(url_for("login"))
        if session.get("user_role") != "ADMIN":
            flash("Apenas administradores podem acessar esta funcionalidade.", "danger")
            return redirect(url_for("index"))
        return funcao(*args, **kwargs)
    return wrapper


def verificar_acesso_site(ocorrencia):
    if session.get("user_role") == "ADMIN":
        return True
    return ocorrencia.site == normalizar_texto(session.get("user_site"))


# =========================
# HELPERS
# =========================
def allowed_image_file(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_IMAGE_EXTENSIONS


def salvar_imagem_upload(file_storage, prefixo="img"):
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None

    nome_original = file_storage.filename.strip()
    if not allowed_image_file(nome_original):
        return None

    ext = nome_original.rsplit(".", 1)[1].lower()
    nome_final = f"{prefixo}_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.{ext}"
    
    # Adicionado fallback caso a função ainda seja usada
    pasta_upload = app.config.get("UPLOAD_FOLDER", "")
    if pasta_upload:
        caminho = os.path.join(pasta_upload, nome_final)
        file_storage.save(caminho)
    return nome_final


# <-- ADICIONADO: Compressor de Imagens Base64 para poupar espaço -->
def processar_imagem_base64(file_storage):
    if not file_storage or not getattr(file_storage, "filename", ""):
        return None
    if not allowed_image_file(file_storage.filename):
        return None
    try:
        img = PILImage.open(file_storage)
        if img.mode != 'RGB':
            img = img.convert('RGB')
        img.thumbnail((800, 800), PILImage.Resampling.LANCZOS)
        buffer = BytesIO()
        img.save(buffer, format="JPEG", quality=70, optimize=True)
        encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
        return f"data:image/jpeg;base64,{encoded}"
    except Exception as e:
        print(f"Erro ao processar imagem: {e}")
        return None


def normalizar_texto(valor: str) -> str:
    return (valor or "").strip().upper()


def normalizar_prioridade(valor: str) -> str:
    valor = normalizar_texto(valor)
    return "CRITICA" if valor == "CRÍTICA" else "MEDIA" if valor == "MÉDIA" else valor


def normalizar_tipo(valor: str) -> str:
    valor = normalizar_texto(valor)
    if valor == "MANUTENÇÃO":
        return "MANUTENCAO"
    if valor == "PENDÊNCIA":
        return "PENDENCIA"
    return valor


def parse_date_or_none(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_datetime_local(value: str):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%dT%H:%M")
    except ValueError:
        return None


def format_date_input(value):
    if not value:
        return ""
    if isinstance(value, str):
        return value
    return value.strftime("%Y-%m-%d")


def format_datetime_local_input(value):
    if not value:
        return ""
    if isinstance(value, str):
        return value
    return value.strftime("%Y-%m-%dT%H:%M")


def pdf_safe(texto):
    if not texto: return "-"
    return escape(str(texto)).replace('\n', '<br/>')


def get_filtros_ocorrencias():
    data_inicial = (request.args.get("data_inicial") or "").strip()
    data_final = (request.args.get("data_final") or "").strip()
    turno = normalizar_texto(request.args.get("turno"))
    status = normalizar_texto(request.args.get("status"))
    
    user_role = session.get("user_role")
    user_site = session.get("user_site")
    
    if user_role != "ADMIN":
        site_filtro = normalizar_texto(user_site)
    else:
        site_filtro = normalizar_texto(request.args.get("site"))

    query = OcorrenciaTurno.query

    if site_filtro:
        query = query.filter(OcorrenciaTurno.site == site_filtro)

    if data_inicial:
        di = parse_date_or_none(data_inicial)
        if di:
            query = query.filter(OcorrenciaTurno.data_ocorrencia >= di)

    if data_final:
        df = parse_date_or_none(data_final)
        if df:
            query = query.filter(OcorrenciaTurno.data_ocorrencia <= df)

    if turno:
        query = query.filter(OcorrenciaTurno.turno == turno)

    if status:
        query = query.filter(OcorrenciaTurno.status == status)

    query = query.order_by(
        OcorrenciaTurno.data_hora_registro.desc(),
        OcorrenciaTurno.id.desc(),
    )

    filtros = {
        "data_inicial": data_inicial,
        "data_final": data_final,
        "turno": turno,
        "status": status,
        "site": site_filtro if user_role == "ADMIN" else "", 
    }
    return query, filtros


def resumo_cards():
    hoje = date.today()
    user_role = session.get("user_role")
    user_site = normalizar_texto(session.get("user_site"))

    def query_base(coluna):
        q = db.session.query(coluna)
        if user_role != "ADMIN":
            q = q.filter(OcorrenciaTurno.site == user_site)
        return q

    ocorrencias_dia = (
        query_base(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.data_ocorrencia == hoje)
        .scalar()
    ) or 0

    pendencias_abertas = (
        query_base(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.status.in_(["EM ABERTO", "EM ACOMPANHAMENTO"]))
        .scalar()
    ) or 0

    turnos_registrados = (
        query_base(func.count(func.distinct(OcorrenciaTurno.turno)))
        .filter(OcorrenciaTurno.data_ocorrencia == hoje)
        .scalar()
    ) or 0

    ocorrencias_criticas = (
        query_base(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.prioridade == "CRITICA")
        .scalar()
    ) or 0

    return {
        "ocorrencias_dia": ocorrencias_dia,
        "pendencias_abertas": pendencias_abertas,
        "turnos_registrados": turnos_registrados,
        "ocorrencias_criticas": ocorrencias_criticas,
    }


def pode_criar_admin_publicamente():
    admin_existe = User.query.filter_by(role="ADMIN").first()
    return admin_existe is None


def ordenar_turnos_query():
    return case(
        (OcorrenciaTurno.turno == "TURNO A", 1),
        (OcorrenciaTurno.turno == "TURNO B", 2),
        (OcorrenciaTurno.turno == "TURNO C", 3),
        (OcorrenciaTurno.turno == "ADM", 4),
        else_=99,
    )


def ordenar_prioridade_query():
    return case(
        (OcorrenciaTurno.prioridade == "CRITICA", 1),
        (OcorrenciaTurno.prioridade == "ALTA", 2),
        (OcorrenciaTurno.prioridade == "MEDIA", 3),
        (OcorrenciaTurno.prioridade == "BAIXA", 4),
        else_=99,
    )


# <-- ALTERADO: Inclusão do fix de buffer (seek 0) e padding para não dar erro no PDF -->
def assinatura_base64_para_image(assinatura_b64, largura_mm=60, altura_mm=22):
    if not assinatura_b64:
        return None

    try:
        if "," in assinatura_b64:
            _, encoded = assinatura_b64.split(",", 1)
        else:
            encoded = assinatura_b64

        encoded += "=" * ((4 - len(encoded) % 4) % 4) 
        image_bytes = base64.b64decode(encoded)
        buffer = BytesIO(image_bytes)
        buffer.seek(0) 
        return Image(buffer, width=largura_mm * mm, height=altura_mm * mm)
    except Exception:
        return None


# <-- ADICIONADO: Helper para desenhar as fotos do Banco direto no PDF -->
def fit_image_b64(base64_str, max_width, max_height):
    if not base64_str:
        return None
    try:
        if "," in base64_str:
            _, encoded = base64_str.split(",", 1)
        else:
            encoded = base64_str
        
        encoded += "=" * ((4 - len(encoded) % 4) % 4) 
        buffer = BytesIO(base64.b64decode(encoded))
        buffer.seek(0) 
        
        img = Image(buffer)
        iw, ih = img.imageWidth, img.imageHeight
        if not iw or not ih:
            return None
        proporcao = min(max_width / float(iw), max_height / float(ih))
        img.drawWidth = iw * proporcao
        img.drawHeight = ih * proporcao
        return img
    except Exception as e:
        print(f"Erro pdf fit_image_b64: {e}")
        return None


def fit_image(path, max_width, max_height):
    try:
        img = Image(path)
        iw, ih = img.imageWidth, img.imageHeight
        if not iw or not ih:
            return None
        proporcao = min(max_width / float(iw), max_height / float(ih))
        img.drawWidth = iw * proporcao
        img.drawHeight = ih * proporcao
        return img
    except Exception:
        return None


# =========================
# ROTAS - AUTH E CONTROLE DE VERSÃO
# =========================
@app.route("/versao-invalida")
def versao_invalida():
    try:
        config_db = SistemaConfig.query.first()
        versao_exigida = config_db.versao_livro if config_db and config_db.versao_livro else "Desconhecida"
    except Exception:
        versao_exigida = "Desconhecida"

    return render_template("versao_invalida.html", versao_local=APP_VERSION, versao_exigida=versao_exigida)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("password") or ""

        user = User.query.filter_by(email=email, is_active=True).first()

        if not user or not user.check_password(senha):
            flash("E-mail ou senha inválidos.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = user.id
        session["username"] = user.nome
        session["user_role"] = user.role
        session["user_site"] = user.site or ""

        flash("Login realizado com sucesso.", "success")
        return redirect(url_for("index"))

    return render_template(
        "login.html",
        permitir_admin_publico=pode_criar_admin_publicamente(),
    )


@app.route("/logout")
def logout():
    session.clear()
    flash("Sessão encerrada com sucesso.", "success")
    return redirect(url_for("login"))


# =========================
# CRIAÇÃO DE USUÁRIO
# =========================
@app.route("/criar-usuario", methods=["GET", "POST"])
def criar_usuario():
    admin_publico_liberado = pode_criar_admin_publicamente()
    usuario_logado_admin = session.get("user_role") == "ADMIN"

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""
        role_solicitado = normalizar_texto(request.form.get("role") or "USER")
        site = request.form.get("site")

        if not nome or not email or not senha:
            flash("Preencha nome, e-mail e senha.", "danger")
            return redirect(url_for("criar_usuario"))

        if role_solicitado not in {"ADMIN", "USER"}:
            role_solicitado = "USER"

        if role_solicitado == "ADMIN" and not (admin_publico_liberado or usuario_logado_admin):
            flash("Somente administradores podem criar outro administrador.", "danger")
            return redirect(url_for("criar_usuario"))

        existe = User.query.filter_by(email=email).first()
        if existe:
            flash("Já existe um usuário com esse e-mail.", "warning")
            return redirect(url_for("criar_usuario"))

        novo = User(
            nome=nome,
            email=email,
            role=role_solicitado,
            site=site,
            is_active=True,
        )
        novo.set_password(senha)

        db.session.add(novo)
        db.session.commit()

        flash("Usuário criado com sucesso.", "success")

        if not session.get("user_id"):
            return redirect(url_for("login"))

        return redirect(url_for("usuarios"))

    sites_db = Site.query.order_by(Site.nome_site.asc()).all()

    return render_template(
        "criar_usuario.html",
        permitir_admin_publico=admin_publico_liberado,
        usuario_logado_admin=usuario_logado_admin,
        sites=sites_db
    )


# =========================
# USUÁRIOS
# =========================
@app.route("/usuarios")
@admin_required
def usuarios():
    lista = User.query.order_by(User.nome.asc()).all()
    # <-- ALTERADO: Correção do Erro 500 no botão de editar usuário -->
    sites_db = Site.query.order_by(Site.nome_site.asc()).all()
    return render_template("usuarios.html", usuarios=lista, sites=sites_db)


@app.route("/usuarios/novo", methods=["GET", "POST"])
@admin_required
def novo_usuario():
    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        senha = request.form.get("senha") or ""
        role = normalizar_texto(request.form.get("role") or "USER")
        site = request.form.get("site")

        if not nome or not email or not senha:
            flash("Preencha nome, e-mail e senha.", "danger")
            return redirect(url_for("novo_usuario"))

        if role not in {"ADMIN", "USER"}:
            role = "USER"

        existe = User.query.filter_by(email=email).first()
        if existe:
            flash("Já existe um usuário com esse e-mail.", "warning")
            return redirect(url_for("novo_usuario"))

        user = User(
            nome=nome,
            email=email,
            role=role,
            site=site,
            is_active=True,
        )
        user.set_password(senha)

        db.session.add(user)
        db.session.commit()

        flash("Usuário criado com sucesso.", "success")
        return redirect(url_for("usuarios"))

    sites_db = Site.query.order_by(Site.nome_site.asc()).all()

    return render_template("usuario_form.html", usuario=None, sites=sites_db)


@app.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
@admin_required
def editar_usuario(user_id):
    usuario = User.query.get_or_404(user_id)

    if request.method == "POST":
        nome = (request.form.get("nome") or "").strip()
        email = (request.form.get("email") or "").strip().lower()
        senha = (request.form.get("senha") or "").strip()
        role = normalizar_texto(request.form.get("role") or "USER")
        site = request.form.get("site")
        is_active = (request.form.get("is_active") or "").strip() == "1"

        if not nome or not email:
            flash("Preencha nome e e-mail.", "danger")
            return redirect(url_for("editar_usuario", user_id=user_id))

        if role not in {"ADMIN", "USER"}:
            role = "USER"

        existe = User.query.filter(User.email == email, User.id != usuario.id).first()
        if existe:
            flash("Já existe outro usuário com esse e-mail.", "warning")
            return redirect(url_for("editar_usuario", user_id=user_id))

        usuario.nome = nome
        usuario.email = email
        usuario.role = role
        usuario.site = site
        usuario.is_active = is_active

        if senha:
            usuario.set_password(senha)

        db.session.commit()

        if usuario.id == session.get("user_id"):
            session["username"] = usuario.nome
            session["user_role"] = usuario.role
            session["user_site"] = usuario.site or ""

        flash("Usuário atualizado com sucesso.", "success")
        return redirect(url_for("usuarios"))

    sites_db = Site.query.order_by(Site.nome_site.asc()).all()

    return render_template("usuario_form.html", usuario=usuario, sites=sites_db)


@app.route("/usuarios/<int:user_id>/excluir", methods=["POST"])
@admin_required
def excluir_usuario(user_id):
    usuario = User.query.get_or_404(user_id)

    if usuario.id == session.get("user_id"):
        flash("Você não pode excluir o próprio usuário logado.", "warning")
        return redirect(url_for("usuarios"))

    db.session.delete(usuario)
    db.session.commit()
    flash("Usuário excluído com sucesso.", "success")
    return redirect(url_for("usuarios"))


# =========================
# OCORRÊNCIAS
# =========================
@app.route("/", methods=["GET"])
@login_required
def index():
    query, filtros = get_filtros_ocorrencias()
    ocorrencias_db = query.all()
    ocorrencias = [o.to_dict() for o in ocorrencias_db]
    ultima_ocorrencia = ocorrencias[0] if ocorrencias else None

    ultimo_id = db.session.query(func.max(OcorrenciaTurno.id)).scalar() or 0
    proximo_id_previsto = ultimo_id + 1

    user_site = session.get("user_site")
    user_id = session.get("user_id")
    if user_site:
        usuarios_mesmo_site = User.query.filter(User.site == user_site, User.id != user_id, User.is_active == True).order_by(User.nome.asc()).all()
    else:
        usuarios_mesmo_site = User.query.filter(User.id != user_id, User.is_active == True).order_by(User.nome.asc()).all()

    sites_db = Site.query.order_by(Site.nome_site.asc()).all()

    return render_template(
        "livro_ocorrencia.html",
        resumo=resumo_cards(),
        ultima_ocorrencia=ultima_ocorrencia,
        ocorrencias=ocorrencias,
        filtros=filtros,
        hoje=date.today().strftime("%Y-%m-%d"),
        proximo_id_previsto=proximo_id_previsto,
        usuarios_mesmo_site=usuarios_mesmo_site,
        sites=sites_db
    )


@app.route("/uploads/<path:filename>")
@login_required
def serve_upload(filename):
    # Rota mantida por segurança caso seja necessário acessar algo antigo, 
    # mas o app usa a pasta app.config (se estiver definida)
    pasta = app.config.get("UPLOAD_FOLDER")
    if pasta and os.path.exists(pasta):
        return send_from_directory(pasta, filename)
    return "Pasta de uploads não configurada.", 404


@app.route("/salvar-ocorrencia-turno", methods=["POST"])
@login_required
def salvar_ocorrencia_turno():
    try:
        data_ocorrencia = parse_date_or_none(request.form.get("data_ocorrencia"))
        data_hora_registro = parse_datetime_local(request.form.get("data_hora_registro"))

        if session.get("user_role") != "ADMIN":
            site = session.get("user_site")
        else:
            site = request.form.get("site")
            
        turno = normalizar_texto(request.form.get("turno"))
        setor = normalizar_texto(request.form.get("setor"))
        tipo_ocorrencia = normalizar_tipo(request.form.get("tipo_ocorrencia"))
        prioridade = normalizar_prioridade(request.form.get("prioridade"))
        
        responsavel_saida = session.get("username", "Usuário")
        responsavel_entrada = (request.form.get("responsavel_entrada") or "").strip()
        
        descricao = (request.form.get("descricao") or "").strip()
        efetivo = (request.form.get("efetivo") or "").strip()

        assinatura_saida = request.form.get("assinatura_saida") or ""
        assinatura_entrada = ""

        acoes_tomadas = (request.form.get("acoes_tomadas") or "").strip()
        pendencias = (request.form.get("pendencias") or "").strip()
        status = normalizar_texto(request.form.get("status"))

        if not all(
            [
                data_ocorrencia,
                data_hora_registro,
                site,
                turno,
                setor,
                tipo_ocorrencia,
                prioridade,
                responsavel_saida,
                responsavel_entrada,
                descricao,
                efetivo,
                status,
                assinatura_saida
            ]
        ):
            flash("Preencha todos os campos obrigatórios e realize sua assinatura.", "danger")
            return redirect(url_for("index"))

        if turno not in TURNOS_VALIDOS:
            flash("Turno inválido.", "danger")
            return redirect(url_for("index"))

        if prioridade not in {"BAIXA", "MEDIA", "ALTA", "CRITICA"}:
            flash("Prioridade inválida.", "danger")
            return redirect(url_for("index"))

        if status not in STATUS_VALIDOS:
            flash("Status inválido.", "danger")
            return redirect(url_for("index"))

        # <-- ALTERADO: Usa o Base64 ao invés de salvar no disco local -->
        imagem_1 = processar_imagem_base64(request.files.get("imagem_1"))
        imagem_2 = processar_imagem_base64(request.files.get("imagem_2"))
        imagem_3 = processar_imagem_base64(request.files.get("imagem_3"))
        imagem_4 = processar_imagem_base64(request.files.get("imagem_4"))

        nova = OcorrenciaTurno(
            data_ocorrencia=data_ocorrencia,
            data_hora_registro=data_hora_registro,
            site=site,
            turno=turno,
            setor=setor,
            tipo_ocorrencia=tipo_ocorrencia,
            prioridade=prioridade,
            responsavel_saida=responsavel_saida,
            responsavel_entrada=responsavel_entrada,
            descricao=descricao,
            efetivo=efetivo,
            assinatura_saida=assinatura_saida,
            assinatura_entrada=None,
            imagem_1=imagem_1,
            imagem_2=imagem_2,
            imagem_3=imagem_3,
            imagem_4=imagem_4,
            acoes_tomadas=acoes_tomadas or None,
            pendencias=pendencias or None,
            status=status,
            criado_por=session.get("username", "Usuário"),
        )

        db.session.add(nova)
        db.session.commit()

        flash("Ocorrência registrada com sucesso. Aguardando assinatura do recebedor.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao salvar ocorrência: {e}", "danger")

    return redirect(url_for("index"))


@app.route("/ocorrencias/<int:ocorrencia_id>/assinar", methods=["GET", "POST"])
@login_required
def assinar_recebimento(ocorrencia_id):
    ocorrencia = OcorrenciaTurno.query.get_or_404(ocorrencia_id)
    
    if not verificar_acesso_site(ocorrencia):
        flash("Acesso negado. Esta ocorrência pertence a outro site.", "danger")
        return redirect(url_for("index"))
    
    if ocorrencia.responsavel_entrada != session.get("username"):
        flash("Você não é o responsável designado para receber este turno.", "danger")
        return redirect(url_for("index"))
        
    if request.method == "POST":
        assinatura = request.form.get("assinatura_entrada")
        if not assinatura:
            flash("A assinatura é obrigatória para receber o turno.", "danger")
            return redirect(url_for("assinar_recebimento", ocorrencia_id=ocorrencia.id))
            
        ocorrencia.assinatura_entrada = assinatura
        ocorrencia.updated_at = datetime.now()
        db.session.commit()
        
        flash("Recebimento de turno assinado com sucesso! Você já pode finalizar a ocorrência.", "success")
        return redirect(url_for("index"))
        
    return render_template("assinar_recebimento.html", ocorrencia=ocorrencia)


@app.route("/ocorrencias/<int:ocorrencia_id>/editar", methods=["GET", "POST"])
@login_required
def editar_ocorrencia(ocorrencia_id):
    ocorrencia = OcorrenciaTurno.query.get_or_404(ocorrencia_id)
    
    if not verificar_acesso_site(ocorrencia):
        flash("Acesso negado. Esta ocorrência pertence a outro site.", "danger")
        return redirect(url_for("index"))

    if request.method == "POST":
        try:
            ocorrencia.data_ocorrencia = parse_date_or_none(request.form.get("data_ocorrencia"))
            ocorrencia.data_hora_registro = parse_datetime_local(
                request.form.get("data_hora_registro")
            )

            if session.get("user_role") == "ADMIN":
                ocorrencia.site = request.form.get("site")
                
            ocorrencia.turno = normalizar_texto(request.form.get("turno"))
            ocorrencia.setor = normalizar_texto(request.form.get("setor"))
            ocorrencia.tipo_ocorrencia = normalizar_tipo(request.form.get("tipo_ocorrencia"))
            ocorrencia.prioridade = normalizar_prioridade(request.form.get("prioridade"))
            ocorrencia.descricao = (request.form.get("descricao") or "").strip()
            ocorrencia.efetivo = (request.form.get("efetivo") or "").strip()

            assinatura_saida_recebida = request.form.get("assinatura_saida") or ""

            if assinatura_saida_recebida:
                ocorrencia.assinatura_saida = assinatura_saida_recebida

            # <-- ALTERADO: Usa o Base64 na atualização de imagens -->
            nova_imagem_1 = processar_imagem_base64(request.files.get("imagem_1"))
            nova_imagem_2 = processar_imagem_base64(request.files.get("imagem_2"))
            nova_imagem_3 = processar_imagem_base64(request.files.get("imagem_3"))
            nova_imagem_4 = processar_imagem_base64(request.files.get("imagem_4"))

            if nova_imagem_1:
                ocorrencia.imagem_1 = nova_imagem_1
            if nova_imagem_2:
                ocorrencia.imagem_2 = nova_imagem_2
            if nova_imagem_3:
                ocorrencia.imagem_3 = nova_imagem_3
            if nova_imagem_4:
                ocorrencia.imagem_4 = nova_imagem_4

            ocorrencia.acoes_tomadas = (request.form.get("acoes_tomadas") or "").strip() or None
            ocorrencia.pendencias = (request.form.get("pendencias") or "").strip() or None
            ocorrencia.status = normalizar_texto(request.form.get("status"))
            ocorrencia.updated_at = datetime.now()

            if not all(
                [
                    ocorrencia.data_ocorrencia,
                    ocorrencia.data_hora_registro,
                    ocorrencia.site,
                    ocorrencia.turno,
                    ocorrencia.setor,
                    ocorrencia.tipo_ocorrencia,
                    ocorrencia.prioridade,
                    ocorrencia.descricao,
                    ocorrencia.efetivo,
                    ocorrencia.status,
                ]
            ):
                flash("Preencha todos os campos obrigatórios.", "danger")
                return redirect(url_for("editar_ocorrencia", ocorrencia_id=ocorrencia.id))

            if ocorrencia.turno not in TURNOS_VALIDOS:
                flash("Turno inválido.", "danger")
                return redirect(url_for("editar_ocorrencia", ocorrencia_id=ocorrencia.id))

            if ocorrencia.prioridade not in {"BAIXA", "MEDIA", "ALTA", "CRITICA"}:
                flash("Prioridade inválida.", "danger")
                return redirect(url_for("editar_ocorrencia", ocorrencia_id=ocorrencia.id))

            if ocorrencia.status not in STATUS_VALIDOS:
                flash("Status inválido.", "danger")
                return redirect(url_for("editar_ocorrencia", ocorrencia_id=ocorrencia.id))

            db.session.commit()
            flash("Ocorrência atualizada com sucesso.", "success")
            return redirect(url_for("index"))

        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao atualizar ocorrência: {e}", "danger")
            return redirect(url_for("editar_ocorrencia", ocorrencia_id=ocorrencia.id))

    sites_db = Site.query.order_by(Site.nome_site.asc()).all()
    return render_template(
        "ocorrencia_form.html",
        ocorrencia=ocorrencia,
        data_ocorrencia_value=format_date_input(ocorrencia.data_ocorrencia),
        data_hora_value=format_datetime_local_input(ocorrencia.data_hora_registro),
        sites=sites_db
    )


@app.route("/ocorrencias/<int:ocorrencia_id>/fechar", methods=["POST"])
@login_required
def fechar_ocorrencia(ocorrencia_id):
    ocorrencia = OcorrenciaTurno.query.get_or_404(ocorrencia_id)
    
    if not verificar_acesso_site(ocorrencia):
        flash("Acesso negado. Esta ocorrência pertence a outro site.", "danger")
        return redirect(url_for("index"))

    if ocorrencia.responsavel_entrada != session.get("username"):
        flash("Apenas o responsável que assumiu o turno pode finalizar esta ocorrência.", "danger")
        return redirect(url_for("index"))

    if not ocorrencia.assinatura_entrada:
        flash("Você precisa assinar o recebimento do turno antes de poder finalizar a ocorrência.", "warning")
        return redirect(url_for("index"))

    try:
        ocorrencia.status = "FINALIZADO"
        ocorrencia.updated_at = datetime.now()
        db.session.commit()
        flash("Ocorrência finalizada com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao finalizar ocorrência: {e}", "danger")

    return redirect(url_for("index"))


@app.route("/ocorrencias/<int:ocorrencia_id>/excluir", methods=["POST"])
@login_required
def excluir_ocorrencia(ocorrencia_id):
    ocorrencia = OcorrenciaTurno.query.get_or_404(ocorrencia_id)
    if not verificar_acesso_site(ocorrencia):
        flash("Acesso negado. Esta ocorrência pertence a outro site.", "danger")
        return redirect(url_for("index"))
        
    try:
        db.session.delete(ocorrencia)
        db.session.commit()
        flash("Ocorrência excluída com sucesso.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Erro ao excluir ocorrência: {e}", "danger")
    return redirect(url_for("index"))


# =========================
# DASHBOARD
# =========================
@app.route("/dashboard-ocorrencias")
@login_required
def dashboard_ocorrencias():
    user_role = session.get("user_role")
    user_site = normalizar_texto(session.get("user_site"))

    def query_dash(coluna):
        q = db.session.query(coluna)
        if user_role != "ADMIN":
            q = q.filter(OcorrenciaTurno.site == user_site)
        return q

    total = query_dash(func.count(OcorrenciaTurno.id)).scalar() or 0
    em_aberto = (
        query_dash(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.status == "EM ABERTO")
        .scalar()
        or 0
    )
    acompanhamento = (
        query_dash(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.status == "EM ACOMPANHAMENTO")
        .scalar()
        or 0
    )
    finalizado = (
        query_dash(func.count(OcorrenciaTurno.id))
        .filter(OcorrenciaTurno.status == "FINALIZADO")
        .scalar()
        or 0
    )

    por_turno = (
        query_dash(func.count(OcorrenciaTurno.id))
        .add_columns(OcorrenciaTurno.turno)
        .group_by(OcorrenciaTurno.turno)
        .order_by(ordenar_turnos_query())
        .all()
    )
    por_turno = [(row[1], row[0]) for row in por_turno]

    por_prioridade = (
        query_dash(func.count(OcorrenciaTurno.id))
        .add_columns(OcorrenciaTurno.prioridade)
        .group_by(OcorrenciaTurno.prioridade)
        .order_by(ordenar_prioridade_query())
        .all()
    )
    por_prioridade = [(row[1], row[0]) for row in por_prioridade]

    por_site = (
        query_dash(func.count(OcorrenciaTurno.id))
        .add_columns(OcorrenciaTurno.site)
        .group_by(OcorrenciaTurno.site)
        .order_by(OcorrenciaTurno.site.asc())
        .all()
    )
    por_site = [(row[1], row[0]) for row in por_site]

    return render_template(
        "dashboard_ocorrencias.html",
        total=total,
        em_aberto=em_aberto,
        acompanhamento=acompanhamento,
        finalizado=finalizado,
        por_turno=por_turno,
        por_prioridade=por_prioridade,
        por_site=por_site,
    )


# =========================
# EXPORTAÇÃO EXCEL
# =========================
@app.route("/export-ocorrencias-excel")
@login_required
def export_ocorrencias_excel():
    query, _ = get_filtros_ocorrencias()
    rows = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Livro de Ocorrências"

    headers = [
        "ID",
        "Data da Ocorrência",
        "Data/Hora Registro",
        "Site",
        "Turno",
        "Setor",
        "Tipo de Ocorrência",
        "Prioridade",
        "Responsável Saída",
        "Responsável Entrada",
        "Efetivo",
        "Descrição",
        "Ações Tomadas",
        "Pendências",
        "Assinatura Saída",
        "Assinatura Entrada",
        "Imagem 1",
        "Imagem 2",
        "Imagem 3",
        "Imagem 4",
        "Status",
        "Criado por",
        "Criado em",
        "Atualizado em",
    ]
    ws.append(headers)

    fill_header = PatternFill("solid", fgColor="FFCC00")
    font_header = Font(bold=True, color="000000")

    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append(
            [
                r.id,
                r.data_ocorrencia.strftime("%d/%m/%Y") if r.data_ocorrencia else "",
                r.data_hora_registro.strftime("%d/%m/%Y %H:%M") if r.data_hora_registro else "",
                r.site,
                r.turno,
                r.setor,
                r.tipo_ocorrencia,
                r.prioridade,
                r.responsavel_saida,
                r.responsavel_entrada,
                r.efetivo or "",
                r.descricao,
                r.acoes_tomadas or "",
                r.pendencias or "",
                "SIM" if r.assinatura_saida else "NÃO",
                "SIM" if r.assinatura_entrada else "NÃO",
                "SIM" if r.imagem_1 else "",
                "SIM" if r.imagem_2 else "",
                "SIM" if r.imagem_3 else "",
                "SIM" if r.imagem_4 else "",
                r.status,
                r.criado_por or "",
                r.created_at.strftime("%d/%m/%Y %H:%M") if r.created_at else "",
                r.updated_at.strftime("%d/%m/%Y %H:%M") if r.updated_at else "",
            ]
        )

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = str(cell.value) if cell.value is not None else ""
            max_length = max(max_length, len(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    nome_arquivo = f"livro_ocorrencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # <-- ALTERADO: O Webview chama a tela "Salvar Como..." do próprio Windows -->
    if len(webview.windows) > 0:
        janela = webview.windows[0]
        destino = janela.create_file_dialog(
            webview.SAVE_DIALOG, 
            save_filename=nome_arquivo, 
            file_types=('Arquivos Excel (*.xlsx)', 'Todos os arquivos (*.*)')
        )
        if destino and len(destino) > 0:
            with open(destino[0], 'wb') as f:
                f.write(output.getvalue())
            flash(f"Planilha Excel salva com sucesso!", "success")
        return redirect(request.referrer or url_for('index'))
    else:
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


# =========================
# EXPORTAÇÃO PDF INDIVIDUAL
# =========================
@app.route("/ocorrencias/<int:ocorrencia_id>/pdf")
@login_required
def export_ocorrencia_individual_pdf(ocorrencia_id):
    ocorrencia = OcorrenciaTurno.query.get_or_404(ocorrencia_id)
    if not verificar_acesso_site(ocorrencia):
        flash("Acesso negado. Esta ocorrência pertence a outro site.", "danger")
        return redirect(url_for("index"))

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()

    titulo_style = ParagraphStyle(
        name="TituloOcorrencia",
        parent=styles["Title"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#d40511"),
        fontSize=16,
        leading=18,
        spaceAfter=4,
    )

    sub_style = ParagraphStyle(
        name="SubOcorrencia",
        parent=styles["Normal"],
        alignment=TA_CENTER,
        textColor=colors.HexColor("#555555"),
        fontSize=8,
        leading=10,
        spaceAfter=6,
    )

    secao_style = ParagraphStyle(
        name="SecaoOcorrencia",
        parent=styles["Heading3"],
        textColor=colors.HexColor("#d40511"),
        fontSize=9,
        leading=10,
        spaceBefore=4,
        spaceAfter=4,
    )

    small_style = ParagraphStyle(
        name="SmallOcorrencia",
        parent=styles["BodyText"],
        fontSize=7,
        leading=8,
        textColor=colors.black,
    )

    def v(valor, default="-"):
        if valor is None:
            return default
        valor = str(valor).strip()
        return pdf_safe(valor) if valor else default

    elements = []
    elements.append(Paragraph("RELATÓRIO INDIVIDUAL DE OCORRÊNCIA", titulo_style))
    elements.append(Paragraph("DHL SECURITY • Passagem de Turno", sub_style))

    faixa_resumo = Table(
        [[
            Paragraph(f"<b>ID:</b> {ocorrencia.id}", small_style),
            Paragraph(f"<b>Status:</b> {v(ocorrencia.status)}", small_style),
            Paragraph(f"<b>Prioridade:</b> {v(ocorrencia.prioridade)}", small_style),
            Paragraph(f"<b>Turno:</b> {v(ocorrencia.turno)}", small_style),
        ]],
        colWidths=[100, 140, 140, 100],
    )
    faixa_resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fff4cc")),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#ffcc00")),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#ffcc00")),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(faixa_resumo)
    elements.append(Spacer(1, 5))

    dados_esquerda = [
        ["ID", str(ocorrencia.id)],
        ["Data da ocorrência", ocorrencia.data_ocorrencia.strftime("%d/%m/%Y") if ocorrencia.data_ocorrencia else "-"],
        ["Data/Hora registro", ocorrencia.data_hora_registro.strftime("%d/%m/%Y %H:%M") if ocorrencia.data_hora_registro else "-"],
        ["Site", ocorrencia.site or "-"],
        ["Turno", ocorrencia.turno or "-"],
        ["Setor", ocorrencia.setor or "-"],
        ["Tipo", ocorrencia.tipo_ocorrencia or "-"],
        ["Prioridade", ocorrencia.prioridade or "-"],
        ["Status", ocorrencia.status or "-"],
    ]

    dados_direita = [
        ["Resp. saída", ocorrencia.responsavel_saida or "-"],
        ["Resp. entrada", ocorrencia.responsavel_entrada or "-"],
        ["Criado por", ocorrencia.criado_por or "-"],
        ["Criado em", ocorrencia.created_at.strftime("%d/%m/%Y %H:%M:%S") if ocorrencia.created_at else "-"],
        ["Atualizado em", ocorrencia.updated_at.strftime("%d/%m/%Y %H:%M:%S") if ocorrencia.updated_at else "-"],
    ]

    tabela_esquerda = Table(dados_esquerda, colWidths=[90, 180])
    tabela_esquerda.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ffcc00")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cfcfcf")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    tabela_direita = Table(dados_direita, colWidths=[100, 170])
    tabela_direita.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ffcc00")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cfcfcf")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    bloco_dados = Table([[tabela_esquerda, tabela_direita]], colWidths=[270, 270])
    bloco_dados.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(bloco_dados)
    elements.append(Spacer(1, 6))

    for titulo, valor in [
        ("Efetivo", ocorrencia.efetivo),
        ("Descrição", ocorrencia.descricao),
        ("Ações tomadas", ocorrencia.acoes_tomadas),
        ("Pendências", ocorrencia.pendencias),
    ]:
        elements.append(Paragraph(titulo, secao_style))
        box = Table([[Paragraph(v(valor), small_style)]], colWidths=[540])
        box.setStyle(TableStyle([
            ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#d9d9d9")),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fbfbfb")),
        ]))
        elements.append(box)
        elements.append(Spacer(1, 5))

    assinatura_saida_img = assinatura_base64_para_image(ocorrencia.assinatura_saida, largura_mm=55, altura_mm=20)
    assinatura_entrada_img = assinatura_base64_para_image(ocorrencia.assinatura_entrada, largura_mm=55, altura_mm=20)

    elements.append(Paragraph("Assinaturas", secao_style))
    assinatura_tabela = Table(
        [[
            assinatura_saida_img if assinatura_saida_img else Paragraph("-", small_style),
            assinatura_entrada_img if assinatura_entrada_img else Paragraph("-", small_style),
        ],
        [
            Paragraph(f"<b>Responsável saída:</b> {v(ocorrencia.responsavel_saida)}", small_style),
            Paragraph(f"<b>Responsável entrada:</b> {v(ocorrencia.responsavel_entrada)}", small_style),
        ]],
        colWidths=[270, 270],
    )
    assinatura_tabela.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#d9d9d9")),
        ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d9d9d9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fcfcfc")),
    ]))
    elements.append(assinatura_tabela)
    elements.append(Spacer(1, 6))

    imagens_registro = [
        ocorrencia.imagem_1,
        ocorrencia.imagem_2,
        ocorrencia.imagem_3,
        ocorrencia.imagem_4,
    ]

    # <-- ALTERADO: O construtor do PDF agora pesquisa as fotos no Banco (Base64) em vez de C:\Uploads -->
    caminhos_validos = [img for img in imagens_registro if img and img.startswith("data:image")]

    if caminhos_validos:
        elements.append(Paragraph("Evidências fotográficas", secao_style))

        qtd = len(caminhos_validos)

        if qtd == 1:
            fotos = []
            for caminho in caminhos_validos:
                img = fit_image_b64(caminho, 500, 180)
                if img:
                    fotos.append([img])
            tabela_fotos = Table(fotos, colWidths=[540], hAlign="CENTER")

        elif qtd == 2:
            linha = []
            for caminho in caminhos_validos:
                img = fit_image_b64(caminho, 250, 160)
                linha.append(img if img else "")
            tabela_fotos = Table([linha], colWidths=[270, 270], hAlign="CENTER")

        else:
            fotos = []
            linha = []
            for caminho in caminhos_validos:
                img = fit_image_b64(caminho, 250, 100)
                linha.append(img if img else "")
                if len(linha) == 2:
                    fotos.append(linha)
                    linha = []

            if linha:
                while len(linha) < 2:
                    linha.append("")
                fotos.append(linha)

            tabela_fotos = Table(fotos, colWidths=[270, 270], hAlign="CENTER")

        tabela_fotos.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#dddddd")),
            ("INNERGRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#dddddd")),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fcfcfc")),
        ]))
        elements.append(tabela_fotos)

    doc.build(elements)
    output.seek(0)

    nome_arquivo = f"ocorrencia_{ocorrencia.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    # <-- ALTERADO: O Webview chama a tela "Salvar Como..." do próprio Windows -->
    if len(webview.windows) > 0:
        janela = webview.windows[0]
        destino = janela.create_file_dialog(
            webview.SAVE_DIALOG, 
            save_filename=nome_arquivo, 
            file_types=('Arquivos PDF (*.pdf)', 'Todos os arquivos (*.*)')
        )
        if destino and len(destino) > 0:
            with open(destino[0], 'wb') as f:
                f.write(output.getvalue())
            flash(f"PDF Individual salvo com sucesso!", "success")
        return redirect(request.referrer or url_for('index'))
    else:
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/pdf",
        )


# =========================
# EXPORTAÇÃO PDF GERAL
# =========================
@app.route("/export-ocorrencias-pdf")
@login_required
def export_ocorrencias_pdf():
    query, filtros = get_filtros_ocorrencias()
    rows = query.all()

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "DHLTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=19,
        leading=22,
        textColor=colors.HexColor("#D40511"),
        alignment=TA_LEFT,
        spaceAfter=4,
    )

    subtitle_style = ParagraphStyle(
        "DHLSubTitle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#444444"),
        alignment=TA_LEFT,
        spaceAfter=8,
    )

    section_label_style = ParagraphStyle(
        "SectionLabel",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=TA_LEFT,
    )

    info_style = ParagraphStyle(
        "InfoStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#333333"),
        alignment=TA_LEFT,
    )

    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.2,
        leading=8.5,
        textColor=colors.HexColor("#1F2937"),
    )

    cell_bold_style = ParagraphStyle(
        "CellBoldStyle",
        parent=cell_style,
        fontName="Helvetica-Bold",
    )

    kpi_value_style = ParagraphStyle(
        "KPIValue",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=18,
        textColor=colors.HexColor("#111111"),
        alignment=TA_CENTER,
    )

    kpi_label_style = ParagraphStyle(
        "KPILabel",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=10,
        textColor=colors.HexColor("#6B7280"),
        alignment=TA_CENTER,
    )

    def valor_seguro(v):
        return pdf_safe(str(v)) if v not in (None, "", "None") else "-"

    def prioridade_badge(prioridade):
        p = (prioridade or "").strip().lower()
        if p in ["alta", "crítica", "critica"]:
            bg = "#FDECEC"
            fg = "#B42318"
        elif p in ["média", "media"]:
            bg = "#FFF4DB"
            fg = "#9A6700"
        elif p in ["baixa"]:
            bg = "#EAF4FF"
            fg = "#175CD3"
        else:
            bg = "#F3F4F6"
            fg = "#374151"

        return Paragraph(
            f"""<para align="center"><font color="{fg}"><b>{valor_seguro(prioridade)}</b></font></para>""",
            ParagraphStyle(
                "badge_prio",
                parent=cell_style,
                backColor=colors.HexColor(bg),
                borderPadding=(3, 5, 3),
                alignment=TA_CENTER,
            ),
        )

    def status_badge(status):
        s = (status or "").strip().lower()
        if s in ["finalizado", "concluída", "concluida", "fechada", "finalizada"]:
            bg = "#E8F7EE"
            fg = "#146C43"
        elif s in ["em acompanhamento", "em andamento", "pendente", "aberta", "em aberto"]:
            bg = "#FFF4DB"
            fg = "#9A6700"
        elif s in ["crítica", "critica", "atrasada"]:
            bg = "#FDE8EA"
            fg = "#B42318"
        else:
            bg = "#EEF2F6"
            fg = "#344054"

        return Paragraph(
            f"""<para align="center"><font color="{fg}"><b>{valor_seguro(status)}</b></font></para>""",
            ParagraphStyle(
                "badge_status",
                parent=cell_style,
                backColor=colors.HexColor(bg),
                borderPadding=(3, 5, 3),
                alignment=TA_CENTER,
            ),
        )

    def p(txt, style=cell_style):
        return Paragraph(valor_seguro(txt), style)

    def tem_sim_nao(valor):
        return "SIM" if valor else "NÃO"

    def total_imagens(r):
        total = 0
        if r.imagem_1:
            total += 1
        if r.imagem_2:
            total += 1
        if r.imagem_3:
            total += 1
        if r.imagem_4:
            total += 1
        return total

    def draw_header_footer(canvas, _doc):
        canvas.saveState()

        page_width, page_height = landscape(A4)

        canvas.setFillColor(colors.HexColor("#D40511"))
        canvas.rect(0, page_height - 12 * mm, page_width, 12 * mm, fill=1, stroke=0)

        canvas.setFillColor(colors.HexColor("#FFCC00"))
        canvas.rect(0, page_height - 14 * mm, page_width, 2 * mm, fill=1, stroke=0)

        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawString(12 * mm, page_height - 8.2 * mm, "DHL SECURITY")

        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(
            page_width - 12 * mm,
            page_height - 8.2 * mm,
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        )

        canvas.setStrokeColor(colors.HexColor("#D1D5DB"))
        canvas.setLineWidth(0.4)
        canvas.line(10 * mm, 10 * mm, page_width - 10 * mm, 10 * mm)

        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#6B7280"))
        canvas.drawString(10 * mm, 6 * mm, "Livro de Ocorrência • Passagem de Turno")
        canvas.drawRightString(page_width - 10 * mm, 6 * mm, f"Página {canvas.getPageNumber()}")

        canvas.restoreState()

    elementos = []

    elementos.append(Paragraph("LIVRO DE OCORRÊNCIAS DE PASSAGEM DE TURNO", title_style))
    elementos.append(
        Paragraph(
            "Relatório corporativo consolidado das ocorrências registradas no sistema.",
            subtitle_style,
        )
    )
    elementos.append(Spacer(1, 4))

    filtros_data = [
        ["Data inicial", valor_seguro(filtros.get("data_inicial"))],
        ["Data final", valor_seguro(filtros.get("data_final"))],
        ["Turno", valor_seguro(filtros.get("turno"))],
        ["Status", valor_seguro(filtros.get("status"))],
        ["Site", valor_seguro(filtros.get("site"))],
    ]

    filtro_table = Table(
        [[Paragraph("<b>FILTROS APLICADOS</b>", section_label_style), ""]]
        + [[Paragraph(f"<b>{k}</b>", info_style), Paragraph(v, info_style)] for k, v in filtros_data],
        colWidths=[45 * mm, 95 * mm],
    )
    filtro_table.setStyle(
        TableStyle(
            [
                ("SPAN", (0, 0), (1, 0)),
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#D40511")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#FAFAFA")),
                ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#D1D5DB")),
                ("INNERGRID", (0, 1), (-1, -1), 0.35, colors.HexColor("#E5E7EB")),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    total_ocorrencias = len(rows)

    kpi_table = Table(
        [
            [Paragraph(str(total_ocorrencias), kpi_value_style)],
            [Paragraph("TOTAL DE OCORRÊNCIAS", kpi_label_style)],
        ],
        colWidths=[42 * mm],
    )

    kpi_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#FFF8DB")),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#FFCC00")),
                ("LEFTPADDING", (0, 0), (-1, -1), 10),
                ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    resumo_bloco = Table([[filtro_table, kpi_table]], colWidths=[145 * mm, 50 * mm])
    resumo_bloco.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )

    elementos.append(KeepTogether(resumo_bloco))
    elementos.append(Spacer(1, 10))

    data = [[
        Paragraph("<b>ID</b>", cell_bold_style),
        Paragraph("<b>DATA / HORA</b>", cell_bold_style),
        Paragraph("<b>SITE</b>", cell_bold_style),
        Paragraph("<b>TURNO</b>", cell_bold_style),
        Paragraph("<b>SETOR</b>", cell_bold_style),
        Paragraph("<b>TIPO</b>", cell_bold_style),
        Paragraph("<b>PRIOR.</b>", cell_bold_style),
        Paragraph("<b>STATUS</b>", cell_bold_style),
        Paragraph("<b>RESP. SAÍDA</b>", cell_bold_style),
        Paragraph("<b>RESP. ENTRADA</b>", cell_bold_style),
        Paragraph("<b>ASS. SAÍDA</b>", cell_bold_style),
        Paragraph("<b>ASS. ENTRADA</b>", cell_bold_style),
        Paragraph("<b>IMGS</b>", cell_bold_style),
    ]]

    for r in rows:
        data.append(
            [
                p(r.id),
                p(r.data_hora_registro.strftime("%d/%m/%Y %H:%M") if r.data_hora_registro else "-"),
                p(r.site),
                p(r.turno),
                p(r.setor),
                p(r.tipo_ocorrencia),
                prioridade_badge(r.prioridade),
                status_badge(r.status),
                p(r.responsavel_saida),
                p(r.responsavel_entrada),
                p(tem_sim_nao(r.assinatura_saida)),
                p(tem_sim_nao(r.assinatura_entrada)),
                p(str(total_imagens(r))),
            ]
        )

    tabela = Table(
        data,
        repeatRows=1,
        colWidths=[
            10 * mm,
            24 * mm,
            18 * mm,
            16 * mm,
            24 * mm,
            34 * mm,
            20 * mm,
            24 * mm,
            30 * mm,
            30 * mm,
            18 * mm,
            18 * mm,
            12 * mm,
        ],
    )

    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFCC00")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FCFCFD")]),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#111827")),
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 7.2),
                ("VALIGN", (0, 1), (-1, -1), "MIDDLE"),
                ("LINEBELOW", (0, 0), (-1, 0), 0.9, colors.HexColor("#D1A800")),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#DADDE1")),
                ("BOX", (0, 0), (-1, -1), 0.7, colors.HexColor("#C9CDD3")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 1), (0, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "CENTER"),
                ("ALIGN", (3, 1), (3, -1), "CENTER"),
                ("ALIGN", (6, 1), (7, -1), "CENTER"),
                ("ALIGN", (10, 1), (12, -1), "CENTER"),
            ]
        )
    )

    elementos.append(tabela)

    doc.build(elementos, onFirstPage=draw_header_footer, onLaterPages=draw_header_footer)

    output.seek(0)
    nome_arquivo = f"livro_ocorrencias_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

    # <-- ALTERADO: O Webview chama a tela "Salvar Como..." do próprio Windows -->
    if len(webview.windows) > 0:
        janela = webview.windows[0]
        destino = janela.create_file_dialog(
            webview.SAVE_DIALOG, 
            save_filename=nome_arquivo, 
            file_types=('Arquivos PDF (*.pdf)', 'Todos os arquivos (*.*)')
        )
        if destino and len(destino) > 0:
            with open(destino[0], 'wb') as f:
                f.write(output.getvalue())
            flash(f"PDF Geral salvo com sucesso!", "success")
        return redirect(request.referrer or url_for('index'))
    else:
        return send_file(
            output,
            as_attachment=True,
            download_name=nome_arquivo,
            mimetype="application/pdf",
        )


# =========================
# SETUP
# =========================
@app.route("/criar-banco")
def criar_banco():
    db.create_all()
    garantir_colunas_ocorrencias()
    garantir_colunas_sistema_config()

    admin = User.query.filter_by(email="admin@dhl.com").first()
    if not admin:
        admin = User(
            nome="Administrador",
            email="admin@dhl.com",
            role="ADMIN",
            site="PG",
            is_active=True,
        )
        admin.set_password("123456")
        db.session.add(admin)
        db.session.commit()

    return "Banco criado com sucesso. Login padrão: admin@dhl.com / 123456"


# =========================
# INICIALIZAÇÃO WEBVIEW
# =========================

# <-- ALTERADO: Iniciar o aplicativo Webview no lugar do servidor comum -->
def start_flask():
    app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        garantir_colunas_ocorrencias()
        garantir_colunas_sistema_config()

        admin = User.query.filter_by(email="admin@dhl.com").first()
        if not admin:
            admin = User(
                nome="Administrador",
                email="admin@dhl.com",
                role="ADMIN",
                site="PG",
                is_active=True,
            )
            admin.set_password("123456")
            db.session.add(admin)
            db.session.commit()

    t = Thread(target=start_flask)
    t.daemon = True
    t.start()

    webview.create_window('Livro de Ocorrências DHL Security', 'http://127.0.0.1:5000', width=1280, height=850, resizable=True)
    webview.start(private_mode=True, debug=True)